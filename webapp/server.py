#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
LogTudo Web Scraper - Web API (FastAPI)
"""

from __future__ import annotations

import json
import os
import threading
import time
import uuid
import traceback
import asyncio
import sys
import unicodedata
import mimetypes
import shutil
import hmac
from contextlib import contextmanager
from datetime import datetime
from dataclasses import dataclass, field
from pathlib import Path
from queue import Queue, Empty
from typing import Any, Dict, List, Optional

from fastapi import FastAPI, File, HTTPException, UploadFile, Request, Depends, Form
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from starlette.middleware.sessions import SessionMiddleware
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse, HTMLResponse, Response, RedirectResponse

from core.automation.login_workflow import LoginWorkflow
from core.automation.nota_fiscal_workflow import NotaFiscalWorkflow
from core.automation.error_handler_workflow import ErrorHandlerWorkflow
from core.automation.playwright_controller import PlaywrightController
from core.services.excel_reader import ExcelReader
from core.services.spreadsheet_writer import SpreadsheetWriter
from core.utils.delay import Delay
from core.utils.error_handler import ErrorHandler
from webapp.db import SessionLocal, Base, engine, DATABASE_URL

try:
    from dotenv import load_dotenv
except Exception:
    load_dotenv = None

if load_dotenv:
    load_dotenv()
from webapp.models import JobRun, JobAction, JobStep, JobError, JobArtifact, BrowserLog

ROOT = Path(__file__).resolve().parent
UPLOAD_DIR = ROOT / "uploads"
EXPORT_DIR = ROOT / "exports"
ARTIFACTS_DIR = EXPORT_DIR / "jobs"

# Configura o diretório de build do frontend (/app/dist no Docker, fallback para webapp/static local)
DIST_DIR = Path("/app/dist")
if not DIST_DIR.exists():
    DIST_DIR = ROOT / "static"

MANUAL_DIR = DIST_DIR / "manual"
ADMIN_DIR = DIST_DIR / "admin"
KNOWN_ERRORS_PATH = Path(__file__).resolve().parent.parent / "config" / "known-errors.json"

BASE_PATH_PLACEHOLDER = "__LOGTUDO_BASE_PATH__"
APP_BASE_PATH = os.getenv("BASE_PATH", "/") or "/"
if not APP_BASE_PATH.startswith("/"):
    APP_BASE_PATH = f"/{APP_BASE_PATH}"
if APP_BASE_PATH != "/" and APP_BASE_PATH.endswith("/"):
    APP_BASE_PATH = APP_BASE_PATH.rstrip("/")
CLIENT_BASE_PATH = "" if APP_BASE_PATH == "/" else APP_BASE_PATH
ADMIN_ROUTE = "/logs-FaturamentosAdicionais"
RESET_LOGS_PASSWORD = os.getenv("LOG_RESET_PASSWORD") or os.getenv("ADMIN_PASS") or ""


def _parse_env_bool(value: Optional[str], default: bool = True) -> bool:
    if value is None:
        return default
    normalized = value.strip().lower()
    if normalized in {"1", "true", "yes", "on"}:
        return True
    if normalized in {"0", "false", "no", "off"}:
        return False
    return default


PLAYWRIGHT_HEADLESS = _parse_env_bool(os.getenv("PLAYWRIGHT_HEADLESS"), default=True)

if sys.platform.startswith("win"):
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())


def _ensure_writable_dir(primary: Path, fallback: Path, label: str) -> Path:
    try:
        primary.mkdir(parents=True, exist_ok=True)
        probe = primary / ".write_test"
        probe.write_text("ok", encoding="utf-8")
        probe.unlink(missing_ok=True)
        return primary
    except Exception as exc:
        print(f"[startup] sem permissao para {label} em {primary}: {exc}")
        fallback.mkdir(parents=True, exist_ok=True)
        probe = fallback / ".write_test"
        probe.write_text("ok", encoding="utf-8")
        probe.unlink(missing_ok=True)
        print(f"[startup] usando fallback de {label}: {fallback}")
        return fallback


UPLOAD_DIR = _ensure_writable_dir(UPLOAD_DIR, Path("/tmp/logtudo/uploads"), "uploads")
EXPORT_DIR = _ensure_writable_dir(EXPORT_DIR, Path("/tmp/logtudo/exports"), "exports")
ARTIFACTS_DIR = _ensure_writable_dir(ARTIFACTS_DIR, EXPORT_DIR / "jobs", "artifacts")
ADMIN_DIR.mkdir(parents=True, exist_ok=True)


def _now_ts() -> str:
    return time.strftime("%d/%m/%Y %H:%M:%S")


def _load_known_errors() -> Dict[str, Any]:
    if not KNOWN_ERRORS_PATH.exists():
        return {}
    try:
        with KNOWN_ERRORS_PATH.open("r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _normalize_text(text: str) -> str:
    if not text:
        return ""
    normalized = unicodedata.normalize("NFKD", text)
    return "".join(c for c in normalized if not unicodedata.combining(c)).lower()


def _match_known_error(message: str, known_errors: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    if not message:
        return None
    normalized_message = _normalize_text(message)
    for key, data in known_errors.items():
        patterns = data.get("match_any", [])
        if any(_normalize_text(p) in normalized_message for p in patterns):
            return {"key": key, "data": data}
    return None


@dataclass
class LogEntry:
    timestamp: str
    level: str
    message: str


@dataclass
class Job:
    id: str
    status: str = "idle"
    current_step: int = 0
    total_steps: int = 0
    current_nf: str = "N/A"
    progress: float = 0.0
    results: List[Dict[str, Any]] = field(default_factory=list)
    logs: List[LogEntry] = field(default_factory=list)
    column_mapping: Dict[str, int] = field(default_factory=dict)
    execute_envios: bool = False
    file_id: str = ""
    file_path: str = ""
    file_source_path: str = ""
    treated_file_path: str = ""
    headers: List[str] = field(default_factory=list)
    data: List[List[Any]] = field(default_factory=list)
    settings: Dict[str, Any] = field(default_factory=dict)
    settings_snapshot: Dict[str, Any] = field(default_factory=dict)
    observability: Dict[str, Any] = field(default_factory=dict)
    human_in_loop: Dict[str, Any] = field(default_factory=dict)
    agent_notes: List[str] = field(default_factory=list)
    is_running: bool = False
    is_paused: bool = False
    stop_requested: bool = False
    log_queue: Queue = field(default_factory=Queue)
    client_ip: str = ""
    user_agent: str = ""
    admin_actor: str = ""

    def add_log(self, level: str, message: str) -> None:
        entry = LogEntry(timestamp=_now_ts(), level=level, message=message)
        self.logs.append(entry)
        self.log_queue.put(entry)


class JobRunner:
    def __init__(self, job: Job):
        self.job = job
        self.delay = Delay()
        self.error_handler = ErrorHandler()
        self.controller: Optional[PlaywrightController] = None
        self.delay.app = self
        self.job_run_id: Optional[str] = None

    @property
    def state(self) -> Dict[str, Any]:
        return {
            "is_running": self.job.is_running,
            "is_paused": self.job.is_paused,
        }

    def log(self, message: str, level: str = "info") -> None:
        level_map = {
            "info": "INFO",
            "success": "SUCCESS",
            "warning": "WARNING",
            "error": "ERROR",
            "debug": "DEBUG",
        }
        self.job.add_log(level_map.get(level, "INFO"), message)

    @contextmanager
    def _db(self):
        session = SessionLocal()
        try:
            yield session
            session.commit()
        except Exception:
            session.rollback()
            raise
        finally:
            session.close()

    def _record_action(self, action_type: str, metadata: Optional[Dict[str, Any]] = None) -> None:
        if not self.job_run_id:
            return
        with self._db() as db:
            db.add(
                JobAction(
                    job_id=self.job_run_id,
                    action_type=action_type,
                    actor=self.job.settings.get("username") or self.job.admin_actor or "user",
                    ip=self.job.client_ip,
                    metadata_json=metadata or {},
                )
            )

    def _record_step_start(self, name: str, metadata: Optional[Dict[str, Any]] = None) -> str:
        if not self.job_run_id:
            return ""
        step_id = str(uuid.uuid4())
        with self._db() as db:
            db.add(
                JobStep(
                    id=step_id,
                    job_id=self.job_run_id,
                    name=name,
                    status="running",
                    metadata_json=metadata or {},
                )
            )
        return step_id

    def _record_step_end(self, step_id: str, status: str = "completed") -> None:
        if not (self.job_run_id and step_id):
            return
        with self._db() as db:
            step = db.get(JobStep, step_id)
            if step:
                step.status = status
                step.ended_at = datetime.utcnow()

    def _record_error(self, message: str, stack: Optional[str] = None, context: Optional[Dict[str, Any]] = None) -> None:
        if not self.job_run_id:
            return
        with self._db() as db:
            db.add(
                JobError(
                    job_id=self.job_run_id,
                    message=message,
                    stack=stack,
                    context=context or {},
                )
            )

    def _record_artifact(self, artifact_type: str, file_path: str) -> None:
        if not self.job_run_id:
            return
        with self._db() as db:
            db.add(
                JobArtifact(
                    job_id=self.job_run_id,
                    type=artifact_type,
                    file_path=file_path,
                )
            )

    def _record_browser_log(self, level: str, message: str, log_type: str, url: Optional[str] = None) -> None:
        if not self.job_run_id:
            return
        with self._db() as db:
            db.add(
                BrowserLog(
                    job_id=self.job_run_id,
                    level=level,
                    message=message,
                    url=url,
                    type=log_type,
                )
            )

    def _record_result_spreadsheet_artifacts(self) -> None:
        if not self.job.file_path:
            return
        try:
            final_path = Path(self.job.file_path).resolve()
            if final_path.exists() and final_path.is_file():
                self._record_artifact("planilha_preenchida", str(final_path))
        except Exception:
            pass

        try:
            uf = str(self.job.settings.get("uf") or "").strip().lower()
            treated_path = str(self.job.treated_file_path or "").strip()
            if uf == "bahia" and treated_path:
                resolved = Path(treated_path).resolve()
                if resolved.exists() and resolved.is_file():
                    self._record_artifact("planilha_tratada_ba", str(resolved))
        except Exception:
            pass

    @contextmanager
    def step(self, name: str, metadata: Optional[Dict[str, Any]] = None):
        step_id = self._record_step_start(name, metadata)
        try:
            yield
            self._record_step_end(step_id, "completed")
        except Exception as e:
            self._record_step_end(step_id, "error")
            self._record_error(str(e), traceback.format_exc(), {"step": name, **(metadata or {})})
            raise

    def _update_progress(self) -> None:
        if self.job.total_steps > 0:
            self.job.progress = round(
                (self.job.current_step / self.job.total_steps) * 100, 2
            )
        else:
            self.job.progress = 0.0

    def _is_cte_filled(self, value: Any) -> bool:
        if value is None:
            return False
        text = str(value).strip()
        if not text:
            return False
        if text.lower() in ("nan", "none"):
            return False
        if text in ("0", "0.0"):
            return False
        return True

    def _execute_treatment(self, treatment: Dict[str, Any]) -> None:
        steps = treatment.get("steps", [])
        for step in steps:
            action = step.get("action")
            value = step.get("value")
            if action == "select_tipo_carga":
                selector = 'select[name="dados_freteMinimo_tipoCarga"]'
                self.controller.page.wait_for_selector(selector, state="visible", timeout=15000)
                self.controller.page.locator(selector).scroll_into_view_if_needed()
                self.delay.custom(int(self.job.settings.get("interaction_delay", 500)))
                self.controller.page.select_option(selector, value)
                self.delay.custom(int(self.job.settings.get("interaction_delay", 500)))
                continue
            raise Exception(f"Tratativa nao suportada: {action}")

    def _try_known_error_recovery(
        self,
        nota_wf: NotaFiscalWorkflow,
        payload: Dict[str, Any],
        error_message: str,
        known_errors: Dict[str, Any],
    ) -> Optional[Dict[str, Any]]:
        match = _match_known_error(error_message, known_errors)
        if not (match and match.get("data", {}).get("treatment")):
            return None

        self.log(
            f"Erro conhecido detectado: {match['key']}. Aplicando tratativa...",
            level="warning",
        )
        self.delay.custom(int(self.job.settings.get("interaction_delay", 500)))
        self._execute_treatment(match["data"]["treatment"])
        self.delay.custom(int(self.job.settings.get("network_delay", 3000)))
        payload["start_from_tag"] = nota_wf.current_tag
        return nota_wf.execute(self.controller.page, payload)

    def run(self) -> None:
        self.job.is_running = True
        self.job.status = "running"
        self.log("Iniciando automacao...", level="info")

        try:
            artifacts_dir = ARTIFACTS_DIR / self.job.id
            artifacts_dir.mkdir(parents=True, exist_ok=True)

            with self._db() as db:
                job_run = JobRun(
                    id=self.job.id,
                    started_at=datetime.utcnow(),
                    status="running",
                    username=self.job.settings.get("username"),
                    ip=self.job.client_ip,
                    user_agent=self.job.user_agent,
                    totals=len(self.job.data),
                    settings_snapshot=self.job.settings_snapshot,
                )
                db.add(job_run)
                self.job_run_id = job_run.id

            self._record_action("start")

            self.controller = PlaywrightController()
            self.controller.start(headless=PLAYWRIGHT_HEADLESS, record_video_dir=str(artifacts_dir))
            self.controller.start_tracing()
            self.log("Navegador iniciado com sucesso", level="success")

            def _on_console(msg):
                try:
                    self._record_browser_log(msg.type, msg.text, "console", self.controller.page.url)
                except Exception:
                    pass

            def _on_page_error(err):
                try:
                    self._record_browser_log("error", str(err), "pageerror", self.controller.page.url)
                except Exception:
                    pass

            def _on_response(resp):
                try:
                    if "/api/jobs" in resp.url:
                        content_type = resp.headers.get("content-type")
                        print("SSE DEBUG:", resp.url, resp.status, content_type)
                        self._record_browser_log(
                            "debug",
                            f"SSE DEBUG: {resp.url} {resp.status} {content_type}",
                            "network",
                            resp.url,
                        )
                    if resp.status >= 400:
                        self._record_browser_log(
                            "error",
                            f"{resp.status} {resp.status_text}",
                            "network",
                            resp.url,
                        )
                except Exception:
                    pass

            self.controller.page.on("console", _on_console)
            self.controller.page.on("pageerror", _on_page_error)
            self.controller.page.on("response", _on_response)

            login_wf = LoginWorkflow(self.delay, self, self.error_handler)
            nota_wf = NotaFiscalWorkflow(self.delay, self, self.error_handler)
            error_workflow = ErrorHandlerWorkflow(self, self.delay)
            self.log("Realizando login...", level="info")
            with self.step("Login"):
                login_wf.execute(self.controller.page, self.job.settings)
            self.log("Login realizado com sucesso", level="success")

            self.job.total_steps = len(self.job.data)
            cte_output_idx = self.job.column_mapping.get("cte_output")
            self.job.current_step = 0
            self.job.current_nf = "N/A"

            filter_expanded = False
            if cte_output_idx is not None:
                processed_count = sum(
                    1
                    for row in self.job.data
                    if cte_output_idx < len(row)
                    and self._is_cte_filled(row[cte_output_idx])
                )
                self.log(
                    f"Detectados {processed_count} registros ja processados na coluna CT-e.",
                    level="info",
                )
            for idx, row in enumerate(self.job.data):
                if not self.job.is_running or self.job.stop_requested:
                    break

                while self.job.is_paused and self.job.is_running:
                    time.sleep(0.5)

                if not self.job.is_running or self.job.stop_requested:
                    break

                self.job.current_step = idx + 1
                self._update_progress()

                if cte_output_idx is not None and cte_output_idx < len(row):
                    val = row[cte_output_idx]
                    if self._is_cte_filled(val):
                        self.log(
                            f"Registro {idx+1} ja processado (CT-e: {str(val).strip()}). Pulando...",
                            level="info",
                        )
                        continue

                nota_fiscal = (
                    str(row[self.job.column_mapping["nota_fiscal"]]).strip()
                    if self.job.column_mapping.get("nota_fiscal") is not None
                    else ""
                )
                tipo_adc = (
                    str(row[self.job.column_mapping["tipo_adc"]]).strip()
                    if self.job.column_mapping.get("tipo_adc") is not None
                    else ""
                )
                valor_cte = (
                    str(row[self.job.column_mapping["valor_cte"]]).strip()
                    if self.job.column_mapping.get("valor_cte") is not None
                    else ""
                )
                senha_ravex = (
                    str(row[self.job.column_mapping["senha_ravex"]]).strip()
                    if self.job.column_mapping.get("senha_ravex") is not None
                    else ""
                )
                transporte = (
                    str(row[self.job.column_mapping["transporte"]]).strip()
                    if self.job.column_mapping.get("transporte") is not None
                    else ""
                )

                self.job.current_nf = nota_fiscal or "N/A"
                self.log(
                    f"Processando registro {idx+1}/{len(self.job.data)} - NF: {nota_fiscal}",
                    level="info",
                )

                try:
                    payload = {
                        "nota_fiscal": nota_fiscal,
                        "tipo_adc": tipo_adc,
                        "valor_cte": valor_cte,
                        "senha_ravex": senha_ravex,
                        "transporte": transporte,
                        "uf": self.job.settings.get("uf"),
                        "should_expand_filter": not filter_expanded,
                        "execute_envios": self.job.execute_envios,
                        "network_delay": int(self.job.settings.get("network_delay", 3000)),
                        "interaction_delay": int(self.job.settings.get("interaction_delay", 500)),
                        "typing_delay": int(self.job.settings.get("typing_delay", 75)),
                    }
                    cte_result = None
                    try:
                        with self.step("Processar NF", {"nf": nota_fiscal, "index": idx + 1}):
                            cte_result = nota_wf.execute(self.controller.page, payload)
                    except Exception as e:
                        self.log(f"Erro ao processar registro: {e}", level="error")
                        error_message = str(e)
                        known_errors = _load_known_errors()
                        match = _match_known_error(error_message, known_errors)
                        if not match:
                            self.log(
                                "Erro nao classificado em known-errors. Usando error handler.",
                                level="warning",
                            )
                        try:
                            cte_result = self._try_known_error_recovery(
                                nota_wf, payload, str(e), known_errors
                            )
                        except Exception as retry_e:
                            self.log(
                                f"Falha ao aplicar tratativa: {retry_e}",
                                level="error",
                            )
                            error_message = str(retry_e)
                            cte_result = None

                        if not cte_result:
                            self.log(
                                "Acionando recuperacao via error handler.",
                                level="warning",
                            )
                            try:
                                error_workflow.handle_recovery(
                                    self.controller.page, login_wf, self.job.settings
                                )
                            except Exception as recovery_e:
                                self.log(
                                    f"Falha no error handler: {recovery_e}",
                                    level="error",
                                )
                                self.job.status = "error"
                                self.job.is_running = False
                                result = {
                                    "status": "error",
                                    "nota_fiscal": nota_fiscal,
                                    "tipo_adc": tipo_adc,
                                    "cte_number": "",
                                    "message": str(recovery_e),
                                    "timestamp": _now_ts(),
                                }
                                self.job.results.append(result)
                                break

                            result = {
                                "status": "error",
                                "nota_fiscal": nota_fiscal,
                                "tipo_adc": tipo_adc,
                                "cte_number": "",
                                "message": error_message,
                                "timestamp": _now_ts(),
                            }
                            self.job.results.append(result)
                            filter_expanded = False
                            continue

                    cte_number = cte_result.get("cte_number", "")
                    result = {
                        "status": "success",
                        "nota_fiscal": nota_fiscal,
                        "tipo_adc": tipo_adc,
                        "cte_number": cte_number,
                        "message": "Processado com sucesso",
                        "timestamp": _now_ts(),
                    }
                    self.job.results.append(result)

                    filter_expanded = True

                    if cte_output_idx is not None:
                        self.job.data[idx][cte_output_idx] = cte_number
                        self._save_spreadsheet_partial()

                    self.log(f"Registro processado. CT-e: {cte_number}", level="success")
                    if not self.job.is_running or self.job.stop_requested:
                        break
                    self.delay.custom(int(self.job.settings.get("step_delay", 1500)))
                except Exception as e:
                    self.log(f"Erro ao processar registro: {e}", level="error")
                    try:
                        if self.controller and self.controller.page:
                            artifacts_dir = ARTIFACTS_DIR / self.job.id
                            artifacts_dir.mkdir(parents=True, exist_ok=True)
                            nf_safe = nota_fiscal.replace("/", "_").replace("\\", "_") if nota_fiscal else "sem_nf"
                            screenshot_path = str((artifacts_dir / f"erro_{idx+1}_{nf_safe}.png").resolve())
                            self.controller.page.screenshot(path=screenshot_path, full_page=True)
                            self._record_artifact("screenshot", screenshot_path)
                    except Exception:
                        pass
                    result = {
                        "status": "error",
                        "nota_fiscal": nota_fiscal,
                        "tipo_adc": tipo_adc,
                        "cte_number": "",
                        "message": str(e),
                        "timestamp": _now_ts(),
                    }
                    self.job.results.append(result)

            if self.job.stop_requested:
                self.job.status = "stopped"
                self.log("Automacao interrompida pelo usuario.", level="warning")
            elif self.job.status != "error":
                self.job.status = "completed"
                self.log("Automacao concluida!", level="success")

        except Exception as e:
            self.job.status = "error"
            self.log(f"Erro critico na automacao: {repr(e)}", level="error")
            self.log(traceback.format_exc(), level="error")
            self._record_error(str(e), traceback.format_exc(), {"stage": "run"})
            try:
                if self.controller and self.controller.page:
                    screenshot_path = str((ARTIFACTS_DIR / self.job.id / "error.png").resolve())
                    self.controller.page.screenshot(path=screenshot_path, full_page=True)
                    self._record_artifact("screenshot", screenshot_path)
            except Exception:
                pass

        finally:
            try:
                if self.controller:
                    trace_path = str((ARTIFACTS_DIR / self.job.id / "trace.zip").resolve())
                    self.controller.stop_tracing(trace_path)
                    self._record_artifact("trace", trace_path)
                    self.controller.stop()
            except Exception:
                pass
            self.job.is_running = False
            self._save_spreadsheet_final()
            self._record_result_spreadsheet_artifacts()
            self._update_progress()
            try:
                artifacts_dir = ARTIFACTS_DIR / self.job.id
                for video in artifacts_dir.glob("**/*.webm"):
                    self._record_artifact("video", str(video.resolve()))
            except Exception:
                pass
            with self._db() as db:
                job_run = db.get(JobRun, self.job.id)
                if job_run:
                    job_run.ended_at = datetime.utcnow()
                    job_run.status = self.job.status
                    total = len(self.job.results)
                    success_count = len([r for r in self.job.results if r.get("status") == "success"])
                    error_count = len([r for r in self.job.results if r.get("status") == "error"])
                    job_run.totals = total
                    job_run.success_count = success_count
                    job_run.error_count = error_count
                    if job_run.started_at and job_run.ended_at:
                        job_run.duration_sec = (job_run.ended_at - job_run.started_at).total_seconds()

    def _save_spreadsheet_partial(self) -> None:
        if not self.job.file_path:
            return
        try:
            from openpyxl import load_workbook

            wb = load_workbook(self.job.file_path)
            ws = wb.active
            cte_output_idx = self.job.column_mapping.get("cte_output")
            if cte_output_idx is None:
                return
            for i, row in enumerate(self.job.data):
                cte_value = row[cte_output_idx]
                if cte_value:
                    ws.cell(row=i + 2, column=cte_output_idx + 1, value=cte_value)
            wb.save(self.job.file_path)
            self.log("Planilha salva automaticamente.", level="info")
        except Exception as e:
            self.log(f"Erro ao salvar planilha automaticamente: {e}", level="error")

    def _save_spreadsheet_final(self) -> None:
        if not self.job.file_path:
            return
        try:
            from openpyxl import load_workbook

            wb = load_workbook(self.job.file_path)
            ws = wb.active
            cte_output_idx = self.job.column_mapping.get("cte_output")
            if cte_output_idx is None:
                return
            for i, row in enumerate(self.job.data):
                cte_value = row[cte_output_idx]
                # Apenas grava o valor se ele existir, para evitar apagar células
                if cte_value:
                    ws.cell(row=i + 2, column=cte_output_idx + 1, value=cte_value)
            wb.save(self.job.file_path)
            self.log("Planilha salva com sucesso.", level="success")
        except Exception as e:
            self.log(f"Erro ao salvar planilha: {e}", level="error")


class JobStore:
    def __init__(self) -> None:
        self.jobs: Dict[str, Job] = {}
        self.files: Dict[str, Dict[str, Any]] = {}

    def create_file(self, file_path: Path, uf: str | None = None) -> Dict[str, Any]:
        reader = ExcelReader()
        data = reader.read(str(file_path), uf=uf)
        selected_path = data.get("file_info", {}).get("full_path", str(file_path))
        source_path = str(file_path)
        treated_path = selected_path if selected_path != source_path else ""
        file_id = str(uuid.uuid4())
        file_payload = {
            "id": file_id,
            "path": selected_path,
            "source_path": source_path,
            "treated_path": treated_path,
            "headers": data["headers"],
            "data": data["data"],
            "file_info": data["file_info"],
            "auto_mapping": reader.auto_map_columns(),
        }
        self.files[file_id] = file_payload
        return file_payload

    def get_file(self, file_id: str) -> Dict[str, Any]:
        if file_id not in self.files:
            raise KeyError("Arquivo nao encontrado")
        return self.files[file_id]

    def create_job(self, payload: Dict[str, Any]) -> Job:
        job_id = str(uuid.uuid4())
        job = Job(id=job_id)
        job.file_id = payload["file_id"]
        file_data = self.get_file(job.file_id)
        job.file_path = file_data["path"]
        job.file_source_path = file_data.get("source_path", "")
        job.treated_file_path = file_data.get("treated_path", "")
        job.headers = file_data["headers"]
        job.data = file_data["data"]
        job.column_mapping = payload["column_mapping"]
        job.execute_envios = payload.get("execute_envios", True)
        job.settings = payload["settings"]
        job.settings_snapshot = payload.get("settings_snapshot", {})
        job.client_ip = payload.get("client_ip", "")
        job.user_agent = payload.get("user_agent", "")
        job.admin_actor = payload.get("admin_actor", "")
        self.jobs[job_id] = job
        return job

    def get_job(self, job_id: str) -> Job:
        if job_id not in self.jobs:
            raise KeyError("Job nao encontrado")
        return self.jobs[job_id]


store = JobStore()
app = FastAPI(title="LogTudo Web API", version="1.0.0", root_path=CLIENT_BASE_PATH)
app.add_middleware(
    SessionMiddleware,
    secret_key=os.getenv("SESSION_SECRET", "change-me"),
    https_only=False,
    same_site="lax",
)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.middleware("http")
async def strip_base_path_middleware(request: Request, call_next):
    path = request.scope.get("path", "")
    print("DEBUG PATH:", path)
    if APP_BASE_PATH != "/" and path.startswith(APP_BASE_PATH):
        request.scope["path"] = path[len(APP_BASE_PATH):]
        print("DEBUG stripped path:", request.scope["path"])
        if not request.scope["path"]:
            request.scope["path"] = "/"
    return await call_next(request)

@app.get("/static/{file_path:path}")
def static_file(file_path: str) -> Response:
    full_path = DIST_DIR / file_path
    if not full_path.exists() or not full_path.is_file():
        raise HTTPException(status_code=404, detail="Arquivo nao encontrado")
    data = full_path.read_bytes()
    media_type, _ = mimetypes.guess_type(str(full_path))
    return Response(content=data, media_type=media_type or "application/octet-stream")


@app.get("/health")
def health_check() -> Dict[str, str]:
    return {"status": "ok"}

def _render_index_html() -> str:
    html = (DIST_DIR / "index.html").read_text(encoding="utf-8")
    html = html.replace(BASE_PATH_PLACEHOLDER, json.dumps(CLIENT_BASE_PATH))
    
    # Força os assets a passarem pela rota explícita /static/ do FastAPI
    # Isso previne que o proxy retorne uma página HTML de erro (Unexpected token '<')
    base = CLIENT_BASE_PATH if CLIENT_BASE_PATH else ""
    html = html.replace('href="./', f'href="{base}/static/')
    html = html.replace('src="./', f'src="{base}/static/')
        
    return html


def _render_admin_html() -> str:
    html = (ADMIN_DIR / "index.html").read_text(encoding="utf-8")
    html = html.replace(BASE_PATH_PLACEHOLDER, json.dumps(CLIENT_BASE_PATH))
    base = CLIENT_BASE_PATH if CLIENT_BASE_PATH else ""
    admin_base = f"{base}{ADMIN_ROUTE}"
    html = html.replace('href="./', f'href="{admin_base}/')
    html = html.replace('src="./', f'src="{admin_base}/')
    return html


@app.get("/")
def index() -> HTMLResponse:
    return HTMLResponse(content=_render_index_html())


@app.get("/index.html")
def index_html() -> HTMLResponse:
    return HTMLResponse(content=_render_index_html())


@app.get("/favicon.ico")
def favicon() -> Response:
    return Response(status_code=204)


@app.get("/manual-usuario")
def manual_index() -> HTMLResponse:
    html = (MANUAL_DIR / "index.html").read_text(encoding="utf-8")
    return HTMLResponse(content=html)


@app.get("/manual-usuario/{file_path:path}")
def manual_file(file_path: str) -> Response:
    full_path = MANUAL_DIR / file_path
    if not full_path.exists() or not full_path.is_file():
        raise HTTPException(status_code=404, detail="Arquivo nao encontrado")
    data = full_path.read_bytes()
    media_type, _ = mimetypes.guess_type(str(full_path))
    return Response(content=data, media_type=media_type or "application/octet-stream")


@app.get(ADMIN_ROUTE)
def admin_index() -> HTMLResponse:
    return HTMLResponse(content=_render_admin_html())


@app.get("/admin")
def admin_legacy_redirect() -> Response:
    return RedirectResponse(url=ADMIN_ROUTE, status_code=307)


@app.get(f"{ADMIN_ROUTE}/{{file_path:path}}")
def admin_file(file_path: str) -> Response:
    if file_path == "index.html":
        return HTMLResponse(content=_render_admin_html())
    full_path = ADMIN_DIR / file_path
    if not full_path.exists() or not full_path.is_file():
        raise HTTPException(status_code=404, detail="Arquivo nao encontrado")
    data = full_path.read_bytes()
    media_type, _ = mimetypes.guess_type(str(full_path))
    return Response(content=data, media_type=media_type or "application/octet-stream")


def _require_admin(_request: Request):
    return None


def _admin_actor(_request: Request) -> str:
    return "admin"


def _client_ip(request: Request) -> str:
    return request.client.host if request.client else ""


def _user_agent(request: Request) -> str:
    return request.headers.get("user-agent", "")


def _resolve_artifact_disk_path(file_path: str, job_id: Optional[str] = None) -> Optional[Path]:
    raw_path = Path(file_path)
    candidates: List[Path] = []

    # 1) Caminho gravado no banco.
    candidates.append(raw_path)

    # 2) Quando o caminho salvo era relativo.
    if not raw_path.is_absolute():
        candidates.append((ROOT / raw_path).resolve())

    # 3) Caminhos alternativos por nome de arquivo dentro da pasta de artefatos do job.
    if job_id and raw_path.name:
        candidates.append((ARTIFACTS_DIR / job_id / raw_path.name).resolve())

    # 4) Compatibilidade com caminhos antigos de container.
    raw_str = str(raw_path).replace("\\", "/")
    legacy_prefixes = ["/app/exports/jobs/", "/app/webapp/exports/jobs/"]
    for prefix in legacy_prefixes:
        if raw_str.startswith(prefix):
            tail = raw_str[len(prefix):].lstrip("/")
            candidates.append((ARTIFACTS_DIR / tail).resolve())

    for candidate in candidates:
        try:
            if candidate.exists() and candidate.is_file():
                return candidate
        except Exception:
            continue
    return None


@app.on_event("startup")
def _startup() -> None:
    if not DATABASE_URL:
        raise RuntimeError("DATABASE_URL nao configurado")
    Base.metadata.create_all(bind=engine)


@app.post("/api/files")
def upload_file(file: UploadFile = File(...), uf: Optional[str] = Form(None)) -> Dict[str, Any]:
    if not file.filename:
        raise HTTPException(status_code=400, detail="Arquivo invalido")
    file_id = str(uuid.uuid4())
    ext = Path(file.filename).suffix.lower()
    if ext not in (".xlsx", ".xls", ".csv"):
        raise HTTPException(status_code=400, detail="Formato nao suportado")
    dest = UPLOAD_DIR / f"{file_id}{ext}"
    with dest.open("wb") as f:
        f.write(file.file.read())
    file_payload = store.create_file(dest, uf=uf)
    if uf and uf.strip().lower() == "bahia":
        treated_path = file_payload.get("file_info", {}).get("full_path")
        if treated_path and str(treated_path) != str(dest):
            print(f"[upload] Bahia detectada: usando arquivo tratado {treated_path}")
    preview = {
        "headers": file_payload["headers"],
        "rows": file_payload["data"][:10],
        "total_rows": len(file_payload["data"]),
        "total_columns": len(file_payload["headers"]),
    }
    return {
        "fileId": file_payload["id"],
        "fileInfo": file_payload["file_info"],
        "preview": preview,
        "autoMapping": file_payload["auto_mapping"],
    }


@app.get("/api/files/{file_id}/preview")
def file_preview(file_id: str) -> Dict[str, Any]:
    try:
        file_data = store.get_file(file_id)
    except KeyError:
        raise HTTPException(status_code=404, detail="Arquivo nao encontrado")
    return {
        "headers": file_data["headers"],
        "rows": file_data["data"][:10],
        "total_rows": len(file_data["data"]),
        "total_columns": len(file_data["headers"]),
    }


@app.post("/api/jobs")
def create_job(request: Request, payload: Dict[str, Any]) -> Dict[str, Any]:
    if "fileId" not in payload or "columnMapping" not in payload or "settings" not in payload:
        raise HTTPException(status_code=400, detail="Payload invalido")
    settings = payload.get("settings") or {}
    if not settings.get("username") or not settings.get("password"):
        raise HTTPException(status_code=400, detail="Credenciais nao configuradas")
    settings_snapshot = {
        **{k: v for k, v in settings.items() if k != "password"},
        "password": "***",
    }

    job = store.create_job(
        {
            "file_id": payload["fileId"],
            "column_mapping": payload["columnMapping"],
            "execute_envios": payload.get("executeEnvios", False),
            "settings": settings,
            "settings_snapshot": settings_snapshot,
            "client_ip": _client_ip(request),
            "user_agent": _user_agent(request),
        }
    )
    runner = JobRunner(job)
    thread = threading.Thread(target=runner.run, daemon=True)
    thread.start()
    return {"jobId": job.id}


@app.get("/api/jobs/{job_id}/status")
def job_status(job_id: str) -> Dict[str, Any]:
    try:
        job = store.get_job(job_id)
    except KeyError:
        raise HTTPException(status_code=404, detail="Job nao encontrado")
    return {
        "status": job.status,
        "progress": job.progress,
        "currentStep": job.current_step,
        "totalSteps": job.total_steps,
        "currentNF": job.current_nf,
        "isPaused": job.is_paused,
    }


@app.post("/api/jobs/{job_id}/pause")
def pause_job(job_id: str) -> Dict[str, Any]:
    job = store.get_job(job_id)
    job.is_paused = True
    job.status = "paused"
    job.add_log("WARNING", "Automacao pausada pelo usuario.")
    with SessionLocal() as db:
        db.add(
            JobAction(
                job_id=job_id,
                action_type="pause",
                actor=job.settings.get("username"),
                ip=job.client_ip,
            )
        )
        db.commit()
    return {"status": "paused"}


@app.post("/api/jobs/{job_id}/resume")
def resume_job(job_id: str) -> Dict[str, Any]:
    job = store.get_job(job_id)
    job.is_paused = False
    job.status = "running"
    job.add_log("INFO", "Automacao retomada.")
    with SessionLocal() as db:
        db.add(
            JobAction(
                job_id=job_id,
                action_type="resume",
                actor=job.settings.get("username"),
                ip=job.client_ip,
            )
        )
        db.commit()
    return {"status": "running"}


@app.post("/api/jobs/{job_id}/stop")
def stop_job(job_id: str) -> Dict[str, Any]:
    job = store.get_job(job_id)
    job.stop_requested = True
    job.is_running = False
    job.status = "stopped"
    job.add_log("WARNING", "Automacao interrompida pelo usuario.")
    with SessionLocal() as db:
        db.add(
            JobAction(
                job_id=job_id,
                action_type="stop",
                actor=job.settings.get("username"),
                ip=job.client_ip,
            )
        )
        db.commit()
    return {"status": "stopped"}


@app.get("/api/jobs/{job_id}/logs")
def job_logs(job_id: str) -> Dict[str, Any]:
    job = store.get_job(job_id)
    return {
        "logs": [
            {"timestamp": l.timestamp, "level": l.level, "message": l.message}
            for l in job.logs
        ]
    }


@app.post("/api/jobs/{job_id}/logs/clear")
def clear_job_logs(job_id: str) -> Dict[str, Any]:
    try:
        job = store.get_job(job_id)
        job.logs.clear()
        return {"status": "ok"}
    except KeyError:
        raise HTTPException(status_code=404, detail="Job nao encontrado")


@app.get("/api/jobs/{job_id}/logs/stream")
def job_logs_stream(job_id: str) -> StreamingResponse:
    try:
        job = store.get_job(job_id)
    except KeyError:
        raise HTTPException(status_code=404, detail="Job nao encontrado")

    def event_stream():
        last_keepalive = time.time()
        while True:
            try:
                entry = job.log_queue.get(timeout=1)
                payload = json.dumps(
                    {
                        "timestamp": entry.timestamp,
                        "level": entry.level,
                        "message": entry.message,
                    },
                    ensure_ascii=False,
                )
                yield f"data: {payload}\n\n"
            except Empty:
                now = time.time()
                if now - last_keepalive > 10:
                    yield "event: ping\ndata: {}\n\n"
                    last_keepalive = now
                if job.status in ("completed", "stopped", "error") and job.log_queue.empty():
                    break

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


@app.get("/api/jobs/{job_id}/results")
def job_results(job_id: str) -> Dict[str, Any]:
    job = store.get_job(job_id)
    return {"results": job.results}


@app.post("/api/jobs/{job_id}/export")
def export_results(job_id: str, payload: Dict[str, Any]) -> FileResponse:
    job = store.get_job(job_id)
    fmt = payload.get("format", "xlsx")
    writer = SpreadsheetWriter()
    export_name = f"resultados_{job_id}.{fmt}"
    export_path = EXPORT_DIR / export_name

    if fmt == "xlsx":
        writer.export_results(job.results, str(export_path))
    elif fmt == "csv":
        writer.export_results_csv(job.results, str(export_path))
    else:
        raise HTTPException(status_code=400, detail="Formato invalido")

    with SessionLocal() as db:
        db.add(
            JobAction(
                job_id=job_id,
                action_type=f"export_{fmt}",
                actor=job.settings.get("username"),
                ip=job.client_ip,
            )
        )
        db.commit()

    return FileResponse(export_path)


RESULT_SPREADSHEET_TYPES = ("planilha_preenchida", "planilha_tratada_ba")


@app.get("/api/results/files")
def list_results_files(limit: int = 100, offset: int = 0) -> Dict[str, Any]:
    with SessionLocal() as db:
        artifacts = (
            db.query(JobArtifact)
            .filter(JobArtifact.type.in_(RESULT_SPREADSHEET_TYPES))
            .order_by(JobArtifact.created_at.desc())
            .offset(max(offset, 0))
            .limit(min(max(limit, 1), 500))
            .all()
        )

        items = []
        for artifact in artifacts:
            job = db.get(JobRun, artifact.job_id) if artifact.job_id else None
            resolved = _resolve_artifact_disk_path(artifact.file_path, artifact.job_id)
            file_name = Path(artifact.file_path).name if artifact.file_path else "arquivo"
            uf = ""
            if job and isinstance(job.settings_snapshot, dict):
                uf = str(job.settings_snapshot.get("uf") or "")

            items.append(
                {
                    "id": artifact.id,
                    "job_id": artifact.job_id,
                    "job_status": job.status if job else "",
                    "job_started_at": job.started_at if job else None,
                    "uf": uf,
                    "type": artifact.type,
                    "file_name": file_name,
                    "file_path": artifact.file_path,
                    "created_at": artifact.created_at,
                    "available": bool(resolved),
                }
            )
        return {"items": items}


@app.get("/api/results/files/{artifact_id}/download")
def download_results_file(artifact_id: str) -> FileResponse:
    with SessionLocal() as db:
        artifact = db.get(JobArtifact, artifact_id)
        if not artifact:
            raise HTTPException(status_code=404, detail="Arquivo nao encontrado")
        if artifact.type not in RESULT_SPREADSHEET_TYPES:
            raise HTTPException(status_code=400, detail="Tipo de arquivo invalido para este endpoint")

    artifact_path = _resolve_artifact_disk_path(artifact.file_path, artifact.job_id)
    if not artifact_path:
        raise HTTPException(status_code=404, detail="Arquivo do resultado nao encontrado")

    media_type, _ = mimetypes.guess_type(str(artifact_path))
    headers = {"Content-Disposition": f'attachment; filename="{artifact_path.name}"'}
    return FileResponse(
        path=str(artifact_path),
        media_type=media_type or "application/octet-stream",
        headers=headers,
    )


@app.post("/api/results/files/delete")
def delete_results_files(payload: Dict[str, Any]) -> Dict[str, Any]:
    if not RESET_LOGS_PASSWORD:
        raise HTTPException(status_code=503, detail="Senha de reset nao configurada no ambiente")

    password = str(payload.get("password") or "")
    if not hmac.compare_digest(password, RESET_LOGS_PASSWORD):
        raise HTTPException(status_code=401, detail="Senha de confirmacao invalida")

    artifact_ids = payload.get("artifact_ids") or []
    if not isinstance(artifact_ids, list) or not artifact_ids:
        raise HTTPException(status_code=400, detail="Nenhum arquivo selecionado")

    if any(job.is_running for job in store.jobs.values()):
        raise HTTPException(status_code=409, detail="Existe automacao em execucao. Pare antes de apagar planilhas.")

    deleted_ids: List[str] = []
    with SessionLocal() as db:
        artifacts = (
            db.query(JobArtifact)
            .filter(JobArtifact.id.in_(artifact_ids), JobArtifact.type.in_(RESULT_SPREADSHEET_TYPES))
            .all()
        )
        if not artifacts:
            return {"status": "ok", "deleted": 0, "deleted_ids": []}

        ids_to_delete = {a.id for a in artifacts}
        paths_to_try: List[tuple[str, Optional[str]]] = []

        for artifact in artifacts:
            paths_to_try.append((artifact.file_path, artifact.job_id))
            deleted_ids.append(artifact.id)
            db.delete(artifact)

        db.commit()

        for file_path, job_id in paths_to_try:
            refs = (
                db.query(JobArtifact)
                .filter(JobArtifact.file_path == file_path)
                .count()
            )
            if refs > 0:
                continue
            resolved = _resolve_artifact_disk_path(file_path, job_id)
            if resolved:
                try:
                    resolved.unlink(missing_ok=True)
                except Exception:
                    pass

    return {"status": "ok", "deleted": len(deleted_ids), "deleted_ids": deleted_ids}


@app.get("/api/admin/summary")
def admin_summary(request: Request) -> Dict[str, Any]:
    _require_admin(request)
    with SessionLocal() as db:
        total_jobs = db.query(JobRun).count()
        success_jobs = db.query(JobRun).filter(JobRun.status == "completed").count()
        error_jobs = db.query(JobRun).filter(JobRun.status == "error").count()
        running_jobs = db.query(JobRun).filter(JobRun.status == "running").count()
        return {
            "total_jobs": total_jobs,
            "success_jobs": success_jobs,
            "error_jobs": error_jobs,
            "running_jobs": running_jobs,
        }


@app.get("/api/admin/jobs")
def admin_jobs(request: Request, limit: int = 50, offset: int = 0, status: Optional[str] = None):
    _require_admin(request)
    with SessionLocal() as db:
        q = db.query(JobRun)
        if status:
            q = q.filter(JobRun.status == status)
        items = (
            q.order_by(JobRun.started_at.desc())
            .offset(offset)
            .limit(min(limit, 200))
            .all()
        )
        return {
            "items": [
                {
                    "id": j.id,
                    "started_at": j.started_at,
                    "ended_at": j.ended_at,
                    "status": j.status,
                    "username": j.username,
                    "ip": j.ip,
                    "totals": j.totals,
                    "success_count": j.success_count,
                    "error_count": j.error_count,
                    "duration_sec": j.duration_sec,
                }
                for j in items
            ]
        }


@app.get("/api/admin/jobs/{job_id}")
def admin_job_detail(request: Request, job_id: str):
    _require_admin(request)
    with SessionLocal() as db:
        j = db.get(JobRun, job_id)
        if not j:
            raise HTTPException(status_code=404, detail="Job nao encontrado")
        return {
            "id": j.id,
            "started_at": j.started_at,
            "ended_at": j.ended_at,
            "status": j.status,
            "username": j.username,
            "ip": j.ip,
            "user_agent": j.user_agent,
            "totals": j.totals,
            "success_count": j.success_count,
            "error_count": j.error_count,
            "duration_sec": j.duration_sec,
            "settings_snapshot": j.settings_snapshot,
        }


@app.get("/api/admin/jobs/{job_id}/actions")
def admin_job_actions(request: Request, job_id: str):
    _require_admin(request)
    with SessionLocal() as db:
        items = (
            db.query(JobAction)
            .filter(JobAction.job_id == job_id)
            .order_by(JobAction.timestamp.desc())
            .all()
        )
        return {
            "items": [
                {
                    "id": a.id,
                    "action_type": a.action_type,
                    "actor": a.actor,
                    "timestamp": a.timestamp,
                    "ip": a.ip,
                    "metadata": a.metadata_json,
                }
                for a in items
            ]
        }


@app.get("/api/admin/jobs/{job_id}/steps")
def admin_job_steps(request: Request, job_id: str):
    _require_admin(request)
    with SessionLocal() as db:
        items = (
            db.query(JobStep)
            .filter(JobStep.job_id == job_id)
            .order_by(JobStep.started_at.asc())
            .all()
        )
        return {
            "items": [
                {
                    "id": s.id,
                    "name": s.name,
                    "status": s.status,
                    "started_at": s.started_at,
                    "ended_at": s.ended_at,
                    "metadata": s.metadata_json,
                }
                for s in items
            ]
        }


@app.get("/api/admin/jobs/{job_id}/artifacts")
def admin_job_artifacts(request: Request, job_id: str):
    _require_admin(request)
    with SessionLocal() as db:
        items = (
            db.query(JobArtifact)
            .filter(JobArtifact.job_id == job_id)
            .order_by(JobArtifact.created_at.desc())
            .all()
        )
        return {
            "items": [
                {
                    "id": a.id,
                    "type": a.type,
                    "file_path": a.file_path,
                    "created_at": a.created_at,
                    "available": bool(_resolve_artifact_disk_path(a.file_path, a.job_id)),
                }
                for a in items
            ]
        }


@app.get("/api/admin/artifacts/{artifact_id}/file")
def admin_artifact_file(request: Request, artifact_id: str):
    _require_admin(request)
    with SessionLocal() as db:
        artifact = db.get(JobArtifact, artifact_id)
        if not artifact:
            raise HTTPException(status_code=404, detail="Artefato nao encontrado")

    artifact_path = _resolve_artifact_disk_path(artifact.file_path, artifact.job_id)
    if not artifact_path:
        raise HTTPException(status_code=404, detail="Arquivo do artefato nao encontrado")

    media_type, _ = mimetypes.guess_type(str(artifact_path))
    headers = {"Content-Disposition": f'inline; filename="{artifact_path.name}"'}
    return FileResponse(
        path=str(artifact_path),
        media_type=media_type or "application/octet-stream",
        headers=headers,
    )


@app.get("/api/admin/jobs/{job_id}/browser-logs")
def admin_job_browser_logs(request: Request, job_id: str):
    _require_admin(request)
    with SessionLocal() as db:
        items = (
            db.query(BrowserLog)
            .filter(BrowserLog.job_id == job_id)
            .order_by(BrowserLog.timestamp.desc())
            .limit(500)
            .all()
        )
        return {
            "items": [
                {
                    "id": l.id,
                    "level": l.level,
                    "message": l.message,
                    "url": l.url,
                    "timestamp": l.timestamp,
                    "type": l.type,
                }
                for l in items
            ]
        }


@app.get("/api/admin/jobs/{job_id}/errors")
def admin_job_errors(request: Request, job_id: str):
    _require_admin(request)
    with SessionLocal() as db:
        items = (
            db.query(JobError)
            .filter(JobError.job_id == job_id)
            .order_by(JobError.timestamp.desc())
            .all()
        )
        return {
            "items": [
                {
                    "id": e.id,
                    "message": e.message,
                    "stack": e.stack,
                    "timestamp": e.timestamp,
                    "context": e.context,
                }
                for e in items
            ]
        }


@app.post("/api/admin/artifacts/{artifact_id}/delete")
def admin_delete_artifact(request: Request, artifact_id: str):
    _require_admin(request)
    with SessionLocal() as db:
        artifact = db.get(JobArtifact, artifact_id)
        if not artifact:
            raise HTTPException(status_code=404, detail="Artefato nao encontrado")
        try:
            Path(artifact.file_path).unlink(missing_ok=True)
        except Exception:
            pass
        db.delete(artifact)
        db.commit()
    return {"status": "ok"}


@app.post("/api/admin/reset-logs")
def admin_reset_logs(request: Request, payload: Dict[str, Any]):
    _require_admin(request)
    if not RESET_LOGS_PASSWORD:
        raise HTTPException(status_code=503, detail="Senha de reset nao configurada no ambiente")

    password = str(payload.get("password") or "")
    if not hmac.compare_digest(password, RESET_LOGS_PASSWORD):
        raise HTTPException(status_code=401, detail="Senha de confirmacao invalida")

    if any(job.is_running for job in store.jobs.values()):
        raise HTTPException(status_code=409, detail="Existe automacao em execucao. Pare antes de resetar os logs.")

    with SessionLocal() as db:
        artifacts = db.query(JobArtifact).all()
        for artifact in artifacts:
            artifact_path = _resolve_artifact_disk_path(artifact.file_path, artifact.job_id)
            if artifact_path:
                try:
                    artifact_path.unlink(missing_ok=True)
                except Exception:
                    pass

        db.query(BrowserLog).delete(synchronize_session=False)
        db.query(JobArtifact).delete(synchronize_session=False)
        db.query(JobError).delete(synchronize_session=False)
        db.query(JobStep).delete(synchronize_session=False)
        db.query(JobAction).delete(synchronize_session=False)
        db.query(JobRun).delete(synchronize_session=False)
        db.commit()

    try:
        if ARTIFACTS_DIR.exists():
            for item in ARTIFACTS_DIR.iterdir():
                if item.is_dir():
                    shutil.rmtree(item, ignore_errors=True)
                else:
                    item.unlink(missing_ok=True)
    except Exception:
        pass

    store.jobs.clear()
    return {"status": "ok"}

# O fallback final: monta toda a pasta dist na raiz para servir assets do frontend (JS, CSS, Imagens, etc.)
# Sem isso, frameworks como Vite/React não conseguem carregar seus arquivos auxiliares e causam tela branca.
if DIST_DIR.exists():
    app.mount("/", StaticFiles(directory=str(DIST_DIR)), name="dist")
