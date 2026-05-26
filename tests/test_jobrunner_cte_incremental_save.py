# -*- coding: utf-8 -*-
"""
Testes para escrita incremental de CTe no JobRunner.
"""

from pathlib import Path

from openpyxl import Workbook, load_workbook

from webapp.server import Job, JobRunner


def _create_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Preview"
    ws.append(["Senha Ravex", "Nota fiscal", "CTe gerado"])
    ws.append(["A1", "NF1", ""])         # alvo para escrita
    ws.append(["A2", "NF2", "123456"])   # já preenchida (não sobrescrever)
    ws.append(["A3", "NF3", ""])         # não deve ser alterada

    ws2 = wb.create_sheet("Dados Extraídos")
    ws2.append(["Senha Ravex", "Nota fiscal", "CTe gerado"])
    ws2.append(["A1", "NF1", ""])
    ws2.append(["A2", "NF2", "123456"])
    ws2.append(["A3", "NF3", ""])
    wb.save(path)


def test_partial_save_updates_only_queued_rows_without_overwrite(tmp_path):
    file_path = tmp_path / "cte.xlsx"
    _create_workbook(file_path)

    job = Job(
        id="job-1",
        file_path=str(file_path),
        headers=["Senha Ravex", "Nota fiscal", "CTe gerado"],
        data=[],
        column_mapping={"senha_ravex": 0, "nota_fiscal": 1, "cte_output": 2},
    )
    runner = JobRunner(job)

    # Enfileira duas atualizações: uma válida vazia e uma que deve ser ignorada por já estar preenchida.
    runner._pending_cte_updates["A1"] = "999001"
    runner._pending_cte_updates["A2"] = "999002"
    runner._save_spreadsheet_partial()

    wb = load_workbook(file_path)
    try:
        for sheet_name in ["Preview", "Dados Extraídos"]:
            ws = wb[sheet_name]
            assert ws.cell(row=2, column=3).value == "999001"
            assert ws.cell(row=3, column=3).value == "123456"
            assert ws.cell(row=4, column=3).value in ("", None)
    finally:
        wb.close()

    assert runner._pending_cte_updates == {}


def test_partial_save_with_no_queue_does_nothing(tmp_path):
    file_path = tmp_path / "cte_noop.xlsx"
    _create_workbook(file_path)

    job = Job(
        id="job-2",
        file_path=str(file_path),
        headers=["Senha Ravex", "Nota fiscal", "CTe gerado"],
        data=[],
        column_mapping={"senha_ravex": 0, "nota_fiscal": 1, "cte_output": 2},
    )
    runner = JobRunner(job)

    runner._save_spreadsheet_partial()

    wb = load_workbook(file_path)
    try:
        for sheet_name in ["Preview", "Dados Extraídos"]:
            ws = wb[sheet_name]
            assert ws.cell(row=2, column=3).value in ("", None)
            assert ws.cell(row=3, column=3).value == "123456"
            assert ws.cell(row=4, column=3).value in ("", None)
    finally:
        wb.close()
