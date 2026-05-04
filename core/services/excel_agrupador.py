import pandas as pd
import warnings
import os
import re
from openpyxl import load_workbook

from core.utils.logger import Logger

# Ignorar avisos do openpyxl sobre estilos padrão do Excel
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# Instancia o logger no padrão do projeto
logger = Logger()

def limpar_texto(texto):
    """Limpa o texto para facilitar a comparação (minúsculas, sem espaços extras)."""
    if pd.isna(texto):
        return ""
    return str(texto).strip().lower()

def limpar_chave(valor):
    """Garante que números de transporte como 12345.0 virem '12345' para não errar o cruzamento."""
    if pd.isna(valor):
        return ""
    txt = str(valor).strip()
    if txt.endswith('.0'):
        txt = txt[:-2]
    return txt


def _ultimo_nao_vazio(series):
    for valor in reversed(list(series.values)):
        if pd.isna(valor):
            continue
        texto = str(valor).strip()
        if not texto or texto.lower() in ["nan", "none"]:
            continue
        return texto[:-2] if texto.endswith(".0") else texto
    return ""


def _todos_nao_vazios(series):
    valores = []
    vistos = set()
    for valor in list(series.values):
        if pd.isna(valor):
            continue
        texto = str(valor).strip()
        if not texto or texto.lower() in ["nan", "none"]:
            continue
        texto = texto[:-2] if texto.endswith(".0") else texto
        chave = texto.lower()
        if chave in vistos:
            continue
        vistos.add(chave)
        valores.append(texto)
    return ", ".join(valores)


def _sheet_dimensions(caminho_arquivo, nome_aba):
    try:
        wb = load_workbook(caminho_arquivo, read_only=True, data_only=True)
        ws = wb[nome_aba]
        dimensions = (ws.max_row or 0, ws.max_column or 0)
        wb.close()
        return dimensions
    except Exception:
        return (0, 0)


def _listar_abas_visiveis(caminho_arquivo):
    wb = load_workbook(caminho_arquivo, read_only=True, data_only=True)
    try:
        return [sheet.title for sheet in wb.worksheets if sheet.sheet_state == "visible"]
    finally:
        wb.close()


def _buscar_dados_zle(caminho_arquivo, nome_aba_zle, transportes_alvo, uf_label):
    if not transportes_alvo:
        return pd.DataFrame(columns=["Chave_Temp", "Valor Frete", "Centro"])

    wb = load_workbook(caminho_arquivo, read_only=True, data_only=True)
    try:
        ws = wb[nome_aba_zle]
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header:
            logger.warning(f"{uf_label}: cabeçalho da ZLE não encontrado.")
            return pd.DataFrame(columns=["Chave_Temp", "Valor Frete", "Centro"])

        col_transp = next((idx for idx, c in enumerate(header) if re.search(r'nº transporte', str(c), re.IGNORECASE)), None)
        col_frete = next((idx for idx, c in enumerate(header) if re.search(r'valor frete', str(c), re.IGNORECASE)), None)
        col_centro = next((idx for idx, c in enumerate(header) if str(c).lower().strip() == 'centro'), None)

        if col_transp is None or col_frete is None or col_centro is None:
            logger.warning(f"{uf_label}: colunas necessárias não encontradas na ZLE.")
            return pd.DataFrame(columns=["Chave_Temp", "Valor Frete", "Centro"])

        encontrados = {}
        total_alvo = len(transportes_alvo)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if col_transp >= len(row):
                continue
            chave = limpar_chave(row[col_transp])
            if not chave or chave not in transportes_alvo or chave in encontrados:
                continue
            encontrados[chave] = {
                "Chave_Temp": chave,
                "Valor Frete": row[col_frete] if col_frete < len(row) else 0,
                "Centro": row[col_centro] if col_centro < len(row) else "Sem Centro",
            }
            if len(encontrados) >= total_alvo:
                break

        logger.success(f"{uf_label}: {len(encontrados)}/{total_alvo} transportes encontrados na ZLE.")
        return pd.DataFrame(encontrados.values(), columns=["Chave_Temp", "Valor Frete", "Centro"])
    finally:
        wb.close()

def processar_planilha_logtudo_agrupada(caminho_arquivo_entrada, caminho_arquivo_saida=None, uf=None):
    uf_label = str(uf).strip() if uf else "UF nao informada"

    if not caminho_arquivo_entrada:
        logger.error(f"{uf_label}: nenhum arquivo de entrada fornecido. Processo cancelado.")
        return None
        
    logger.info(f"{uf_label}: iniciando o processamento e agrupamento do arquivo: {caminho_arquivo_entrada}")
    
    if not caminho_arquivo_saida:
        diretorio = os.path.dirname(caminho_arquivo_entrada)
        nome_arquivo = os.path.basename(caminho_arquivo_entrada)
        caminho_arquivo_saida = os.path.join(diretorio, f"Processado_Agrupado_{nome_arquivo}")

    mapa_colunas = {
        'ID': ['id', 'senha', 'senha ravex', 'id ravex'],
        'Tipo de custo': ['tipo de custo', 'tipo adc', 'tipo do custo'],
        'Nota fiscal': ['nota fiscal', 'nf'],
        'Transporte': ['transporte', 'transporte adc', 'transporte adc criado', 'transporte criado']
    }

    try:
        abas_visiveis = _listar_abas_visiveis(caminho_arquivo_entrada)
        
        # 1. Identificar a aba alvo ("Logtudo" ou "Base")
        nome_aba_base = None
        nome_aba_zle = None
        nome_aba_valores = None
        
        for nome_aba in abas_visiveis:
            nome_limpo = str(nome_aba).lower()
            if 'logtudo' in nome_limpo or 'base' in nome_limpo:
                nome_aba_base = nome_aba
            elif 'zle' in nome_limpo:
                nome_aba_zle = nome_aba
            elif 'valores' in nome_limpo:
                nome_aba_valores = nome_aba
                
        if not nome_aba_base:
            logger.error(f"{uf_label}: nenhuma aba contendo 'Logtudo' ou 'Base' foi encontrada.")
            return None

        logger.success(f"{uf_label}: aba Base identificada: '{nome_aba_base}'")
        df_bruto = pd.read_excel(caminho_arquivo_entrada, sheet_name=nome_aba_base, header=None)

        # 2. Localizar cabeçalho da Base
        indice_cabecalho = -1
        mapeamento_indices = {'ID': -1, 'Tipo de custo': -1, 'Nota fiscal': -1, 'Transporte': -1}
        
        for i in range(min(10, len(df_bruto))):
            linha = df_bruto.iloc[i].tolist()
            matches = 0
            indices_temporarios = {'ID': -1, 'Tipo de custo': -1, 'Nota fiscal': -1, 'Transporte': -1}
            
            for col_idx, valor_celula in enumerate(linha):
                valor_limpo = limpar_texto(valor_celula)
                for chave_padrao, variacoes in mapa_colunas.items():
                    if valor_limpo in variacoes:
                        indices_temporarios[chave_padrao] = col_idx
                        matches += 1
                        break
            
            if matches >= 2:
                indice_cabecalho = i
                mapeamento_indices = indices_temporarios
                logger.success(f"{uf_label}: cabeçalho encontrado na linha {i + 1} do Excel.")
                break
                
        if indice_cabecalho == -1:
            logger.error(f"{uf_label}: não foi possível identificar o cabeçalho na aba Base.")
            return None

        # 3. Extrair os dados da Base
        logger.info(f"{uf_label}: extraindo dados da Base...")
        df_dados = pd.read_excel(caminho_arquivo_entrada, sheet_name=nome_aba_base, header=indice_cabecalho)
        
        colunas_para_extrair = {df_dados.columns[idx]: chave for chave, idx in mapeamento_indices.items() if idx != -1}
        df_extraido = df_dados[list(colunas_para_extrair.keys())].copy()
        df_extraido.rename(columns=colunas_para_extrair, inplace=True)
        df_extraido.dropna(how='all', inplace=True)
        
        # 4. BUSCAR VALORES DA ABA ZLE
        if nome_aba_zle and 'Transporte' in df_extraido.columns:
            logger.info(f"{uf_label}: cruzando dados com a aba '{nome_aba_zle}'...")
            df_extraido['Chave_Temp'] = df_extraido['Transporte'].apply(limpar_chave)
            transportes_alvo = set(df_extraido['Chave_Temp'].dropna().astype(str))
            transportes_alvo.discard("")
            df_zle_subset = _buscar_dados_zle(caminho_arquivo_entrada, nome_aba_zle, transportes_alvo, uf_label)

            if not df_zle_subset.empty:
                df_extraido = pd.merge(df_extraido, df_zle_subset, on='Chave_Temp', how='left')
                df_extraido.drop('Chave_Temp', axis=1, inplace=True)
                logger.success(f"{uf_label}: colunas 'Valor Frete' e 'Centro' adicionadas com sucesso.")
            else:
                df_extraido.drop('Chave_Temp', axis=1, inplace=True)
                logger.warning(f"{uf_label}: nenhum dado da ZLE foi cruzado.")
        
        # Garantir que as colunas existam mesmo se a ZLE falhou
        if 'Centro' not in df_extraido.columns: df_extraido['Centro'] = 'Sem Centro'
        if 'Valor Frete' not in df_extraido.columns: df_extraido['Valor Frete'] = 0

        # ---------------------------------------------------------
        # 5. NOVA FUNCIONALIDADE: TRATAMENTO E AGRUPAMENTO
        # ---------------------------------------------------------
        logger.info(f"{uf_label}: tratando nomes das colunas e agrupando tabelas...")
        
        # Renomear colunas
        df_extraido.rename(columns={
            'ID': 'Senha Ravex',
            'Transporte': 'Nº Transporte',
            'Centro': 'Tipo Cte'
        }, inplace=True)
        
        # Criar a nova coluna CTe gerado vazia
        df_extraido['CTe gerado'] = ""
        
        # Garantir que Valor Frete é numérico para conseguir somar depois
        df_extraido['Valor Frete'] = pd.to_numeric(df_extraido['Valor Frete'], errors='coerce').fillna(0)
        
        # Reordenar colunas para ficar visualmente bonito
        ordem_colunas = ['Senha Ravex', 'Tipo de custo', 'Nota fiscal', 'Nº Transporte', 'Valor Frete', 'Tipo Cte', 'CTe gerado']
        df_extraido = df_extraido[[c for c in ordem_colunas if c in df_extraido.columns]]

        lista_relatorio = []
        linhas_operacionais = []
        
        # Lista para armazenar os dados de cada tabela (para uso futuro)
        tabelas_data = []
        
        # Pega todos os tipos de Cte únicos (removendo vazios e normalizando)
        df_extraido['Tipo Cte'] = df_extraido['Tipo Cte'].fillna('NÃO IDENTIFICADO')
        tipos_unicos = df_extraido['Tipo Cte'].unique()

        for tipo in tipos_unicos:
            # Filtra os dados deste "Tipo Cte" específico
            df_tipo = df_extraido[df_extraido['Tipo Cte'] == tipo].copy()
            
            if df_tipo.empty:
                continue
                
            # Obter os valores consolidados da linha operacional, seguindo o modelo antigo da BA.
            senha_resumo = _todos_nao_vazios(df_tipo['Senha Ravex'])
            nf_resumo = _todos_nao_vazios(df_tipo['Nota fiscal'])
            transporte_resumo = _todos_nao_vazios(df_tipo['Nº Transporte'])
            ultimo_tipo_custo = _ultimo_nao_vazio(df_tipo['Tipo de custo'])
            soma_frete = df_tipo['Valor Frete'].sum()
                
            # Coletar os arrays de dados para cada tabela
            tabelas_data.append({
                'tipo_custo': ultimo_tipo_custo,  # Último tipo de custo
                'senha_ravex': df_tipo['Senha Ravex'].tolist(),
                'nota_fiscal': df_tipo['Nota fiscal'].tolist(),
                'transporte': df_tipo['Nº Transporte'].tolist(),
                'valor_cte': soma_frete  # Adicionar valor total
            })
            
            # Montar a linha de resumo (Dicionário com as colunas)
            resumo = {
                'Senha Ravex': senha_resumo,
                'Tipo de custo': ultimo_tipo_custo,
                'Nota fiscal': nf_resumo,
                'Nº Transporte': transporte_resumo,
                'Valor Frete': soma_frete,
                'Tipo Cte': tipo,
                'CTe gerado': ''
            }
            resumo_relatorio = resumo.copy()
            resumo_relatorio['Tipo Cte'] = 'RESUMO'
            linha_resumo = pd.DataFrame([resumo_relatorio])
            linhas_operacionais.append(resumo)
            
            # Linha em branco para separar as tabelas
            linha_branca = pd.DataFrame([{col: '' for col in df_extraido.columns}])
            
            # Empilhar: Tabela do tipo -> Resumo -> Linha Branca
            lista_relatorio.append(df_tipo)
            lista_relatorio.append(linha_resumo)
            lista_relatorio.append(linha_branca)

        # Junta tudo de volta em um único DataFrame formatado
        if lista_relatorio:
            df_final = pd.concat(lista_relatorio, ignore_index=True)
            df_operacional = pd.DataFrame(linhas_operacionais, columns=df_extraido.columns)
            logger.success(f"{uf_label}: tratamento e agrupamento concluídos.")
        else:
            logger.warning(f"{uf_label}: nenhum dado válido para agrupar.")
            df_final = df_extraido
            df_operacional = df_extraido

        # ---------------------------------------------------------

        # 6. Salvar o novo arquivo Excel
        logger.info(f"{uf_label}: gerando arquivo final: {caminho_arquivo_saida}")
        with pd.ExcelWriter(caminho_arquivo_saida, engine='openpyxl') as writer:
            # Salva aba base cortando nome se muito grande pro Excel (máx 31 chars)
            nome_base_sheet = f"Base ({str(nome_aba_base)[:20]})"
            df_bruto.to_excel(writer, sheet_name=nome_base_sheet, index=False, header=False)
            
            # Aba usada pela automação: uma linha consolidada por bloco.
            df_operacional.to_excel(writer, sheet_name='Dados Extraídos', index=False)

            # Aba visual com detalhes, resumo e separadores.
            df_final.to_excel(writer, sheet_name='Relatório Agrupado', index=False)
            
            if nome_aba_valores:
                df_valores = pd.read_excel(caminho_arquivo_entrada, sheet_name=nome_aba_valores, header=None)
                df_valores.to_excel(writer, sheet_name=str(nome_aba_valores)[:31], index=False, header=False)

            if nome_aba_zle:
                zle_rows, _ = _sheet_dimensions(caminho_arquivo_entrada, nome_aba_zle)
                if zle_rows > 50000:
                    logger.warning(
                        f"{uf_label}: aba '{nome_aba_zle}' possui {zle_rows} linhas e não foi copiada para o arquivo tratado."
                    )
                else:
                    df_zle_original = pd.read_excel(caminho_arquivo_entrada, sheet_name=nome_aba_zle, header=None)
                    df_zle_original.to_excel(writer, sheet_name=str(nome_aba_zle)[:31], index=False, header=False)

        logger.success(f"{uf_label}: processo concluído com sucesso!")
        return caminho_arquivo_saida, tabelas_data

    except Exception as e:
        logger.error(f"{uf_label}: ocorreu um erro durante a execução: {e}")
        return None
