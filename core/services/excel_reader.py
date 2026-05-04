# -*- coding: utf-8 -*-
"""
ExcelReader - Módulo de Leitura de Excel
Lê arquivos Excel (xlsx, xls, csv) e extrai dados
"""

from pathlib import Path
import pandas as pd
import os
import unicodedata


class ExcelReader:
    """Classe para leitura de arquivos Excel e CSV"""

    def __init__(self):
        """Inicializa o leitor de Excel"""
        self.headers = []
        self.data = []
        self.file_info = None

    def read(self, file_path, uf=None):
        """
        Lê arquivo de planilha

        Args:
            file_path: Caminho do arquivo
            uf: UF selecionada (opcional)

        Returns:
            Dicionário com headers, data e file_info

        Raises:
            Exception: Se houver erro ao processar o arquivo
        """
        file_path = str(file_path)
        extension = self._get_file_extension(file_path)

        if extension == 'csv':
            return self._read_csv(file_path)
        elif extension in ['xlsx', 'xls']:
            if self._uses_grouped_treatment(uf):
                return self._process_grouped(file_path, uf)
            return self._read_excel(file_path, extension)
        else:
            raise Exception(f"Formato de arquivo não suportado: {extension}")

    def _get_file_extension(self, filename):
        """Obtém extensão do arquivo"""
        parts = filename.split('.')
        return parts[-1].lower() if parts else ''

    def _normalize_uf(self, uf):
        """Normaliza a UF selecionada para comparações internas."""
        text = str(uf or "").strip()
        normalized = unicodedata.normalize("NFKD", text)
        normalized = "".join(c for c in normalized if not unicodedata.combining(c))
        return normalized.lower()

    def _uses_grouped_treatment(self, uf):
        """Indica se a UF deve usar o modelo novo agrupado."""
        return self._normalize_uf(uf) in {"bahia", "ba", "pernambuco", "pe", "ceara", "ce"}

    def _process_grouped(self, file_path, uf=None):
        """
        Processa arquivo usando o modelo agrupado para BA, PE e CE.
        Gera um arquivo tratado e lê a aba Dados Extraídos.
        """
        try:
            from core.services.excel_agrupador import processar_planilha_logtudo_agrupada
            
            path = Path(file_path)

            if self._is_grouped_treated_workbook(path):
                return self._read_excel(str(path), 'xlsx', sheet_name='Dados Extraídos')

            # Se não estiver totalmente tratada, identificar blocos pendentes
            pending_blocks = self.get_pending_blocks(str(path))
            if pending_blocks:
                print(f"Encontrados {len(pending_blocks)} blocos pendentes para processamento")
                # Aqui você pode adicionar lógica adicional, como logging ou notificação

            new_filename = f"Processado_Agrupado_{path.stem}.xlsx"
            new_path = path.parent / new_filename
            
            result = processar_planilha_logtudo_agrupada(file_path, str(new_path), uf=uf)
            if not result:
                raise Exception("O agrupador não retornou arquivo tratado")

            if isinstance(result, tuple):
                treated_path = result[0]
            else:
                treated_path = result
            
            return self._read_excel(str(treated_path), 'xlsx', sheet_name='Dados Extraídos')
            
        except Exception as e:
            uf_label = str(uf or "UF selecionada").strip()
            raise Exception(f"Erro ao processar planilha de {uf_label}: {str(e)}")

    def _is_grouped_treated_workbook(self, file_path):
        """Verifica se o arquivo ja e uma planilha agrupada pronta para automacao."""
        try:
            from openpyxl import load_workbook

            wb = load_workbook(file_path, read_only=True, data_only=True)
            try:
                if 'Dados Extraídos' not in wb.sheetnames:
                    return False
                ws = wb['Dados Extraídos']
                header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                if not header:
                    return False
                normalized_headers = {self._normalize_header(h) for h in header if h is not None}
                required_headers = {
                    self._normalize_header('Senha Ravex'),
                    self._normalize_header('Tipo de custo'),
                    self._normalize_header('Nota fiscal'),
                    self._normalize_header('Nº Transporte'),
                    self._normalize_header('Valor Frete'),
                    self._normalize_header('CTe gerado'),
                }
                if not required_headers.issubset(normalized_headers):
                    return False
                
                # Verificar se há pelo menos um CTe gerado preenchido
                cte_col_idx = None
                for idx, h in enumerate(header):
                    if self._normalize_header(h) == self._normalize_header('CTe gerado'):
                        cte_col_idx = idx
                        break
                
                if cte_col_idx is None:
                    return False
                
                has_cte_generated = False
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[cte_col_idx] and str(row[cte_col_idx]).strip():
                        has_cte_generated = True
                        break
                
                return has_cte_generated
            finally:
                wb.close()
        except Exception:
            return False

    def get_pending_blocks(self, file_path):
        """
        Identifica blocos pendentes na aba 'Relatório agrupado' que ainda precisam ser processados.
        
        Args:
            file_path: Caminho do arquivo Excel
            
        Returns:
            Lista de dicionários com informações dos blocos pendentes
        """
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(file_path, read_only=True, data_only=True)
            try:
                if 'Relatório agrupado' not in wb.sheetnames:
                    return []
                
                ws = wb['Relatório agrupado']
                header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                if not header:
                    return []
                
                # Normalizar headers para busca
                normalized_headers = [self._normalize_header(h) for h in header]
                
                # Encontrar índices das colunas relevantes
                senha_idx = None
                tipo_idx = None
                nf_idx = None
                transporte_idx = None
                valor_idx = None
                cte_idx = None
                
                for idx, h in enumerate(normalized_headers):
                    if h == self._normalize_header('Senha Ravex'):
                        senha_idx = idx
                    elif h == self._normalize_header('Tipo de custo'):
                        tipo_idx = idx
                    elif h == self._normalize_header('Nota fiscal'):
                        nf_idx = idx
                    elif h == self._normalize_header('Nº Transporte'):
                        transporte_idx = idx
                    elif h == self._normalize_header('Valor Frete'):
                        valor_idx = idx
                    elif h == self._normalize_header('CTe gerado'):
                        cte_idx = idx
                
                pending_blocks = []
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    # Verificar se CTe gerado está vazio
                    cte_value = row[cte_idx] if cte_idx is not None else None
                    if not cte_value or not str(cte_value).strip():
                        # Bloco pendente encontrado
                        block_info = {
                            'senha_ravex': row[senha_idx] if senha_idx is not None else '',
                            'tipo_custo': row[tipo_idx] if tipo_idx is not None else '',
                            'nota_fiscal': row[nf_idx] if nf_idx is not None else '',
                            'transporte': row[transporte_idx] if transporte_idx is not None else '',
                            'valor_frete': row[valor_idx] if valor_idx is not None else '',
                            'cte_gerado': cte_value or ''
                        }
                        pending_blocks.append(block_info)
                
                return pending_blocks
                
            finally:
                wb.close()
        except Exception as e:
            print(f"Erro ao identificar blocos pendentes: {str(e)}")
            return []

    def _normalize_header(self, value):
        text = str(value or "").strip()
        normalized = unicodedata.normalize("NFKD", text)
        normalized = "".join(c for c in normalized if not unicodedata.combining(c))
        return normalized.lower()

    def _read_excel(self, file_path, extension, sheet_name=None):
        """
        Lê arquivo Excel usando openpyxl para preservar formatação

        Args:
            file_path: Caminho do arquivo
            extension: Extensão do arquivo

        Returns:
            Dicionário com dados processados
        """
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(file_path)
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    raise Exception(f"Aba '{sheet_name}' não encontrada")
                ws = wb[sheet_name]
            else:
                ws = wb.active
            
            # Ler headers da primeira linha
            self.headers = []
            for cell in ws[1]:
                self.headers.append(cell.value if cell.value is not None else '')
            
            # Ler dados a partir da segunda linha
            self.data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                self.data.append(list(row))
            
            # Converter valores None para strings vazias
            self.data = [['' if v is None else v for v in row] for row in self.data]

            # Obter informações do arquivo
            file_size = os.path.getsize(file_path)

            self.file_info = {
                'name': Path(file_path).name,
                'size': file_size,
                'extension': extension,
                'rows': len(self.data),
                'columns': len(self.headers),
                'full_path': file_path
            }

            return {
                'headers': self.headers,
                'data': self.data,
                'file_info': self.file_info
            }

        except Exception as e:
            raise Exception(f"Erro ao processar Excel: {str(e)}")

    def _read_csv(self, file_path):
        """
        Lê arquivo CSV

        Args:
            file_path: Caminho do arquivo

        Returns:
            Dicionário com dados processados
        """
        try:
            # Ler arquivo CSV
            df = pd.read_csv(file_path, encoding='utf-8')

            # Pular linhas com 'Nº CTE' preenchido
            cte_col_names = ['Nº CTE', 'nº cte', 'numero cte', 'número cte', 'cte', 'n_cte', 'nºcte']
            cte_column = None
            for col_name in df.columns:
                if col_name.strip().lower() in [c.lower() for c in cte_col_names]:
                    cte_column = col_name
                    break
            
            if cte_column:
                df = df[pd.isna(df[cte_column]) | (df[cte_column] == '')]

            # Converter para listas
            self.headers = df.columns.tolist()
            self.data = df.values.tolist()

            # Converter valores None para strings vazias
            self.data = [['' if v is None else v for v in row] for row in self.data]

            # Obter informações do arquivo
            file_size = os.path.getsize(file_path)

            self.file_info = {
                'name': Path(file_path).name,
                'size': file_size,
                'extension': 'csv',
                'rows': len(self.data),
                'columns': len(self.headers)
            }

            return {
                'headers': self.headers,
                'data': self.data,
                'file_info': self.file_info
            }

        except Exception as e:
            raise Exception(f"Erro ao processar CSV: {str(e)}")

    def find_column_by_name(self, column_names):
        """
        Encontra coluna pelo nome

        Args:
            column_names: Lista de nomes possíveis para a coluna

        Returns:
            Índice da coluna ou -1 se não encontrada
        """
        for name in column_names:
            normalized_name = name.lower().strip()
            for i, header in enumerate(self.headers):
                if header.lower().strip() == normalized_name:
                    return i
        return -1

    def auto_map_columns(self):
        """
        Mapeia colunas automaticamente

        Returns:
            Dicionário com mapeamento encontrado
        """
        mappings = {
            'nota_fiscal': ['Nota fiscal', 'NOTA FISCAL', 'nota fiscal', 'nf', 'nº nf'],
            'tipo_adc': ['Tipo de custo', 'TIPO ADC', 'tipo adc', 'tipoadc', 'adc', 'tipo adicional', 'tipo'],
            'valor_cte': ['Valor Frete', 'VALOR TT CTE', 'valor tt cte', 'valor_tt_cte', 'valor cte', 'valor_cte', 'valor', 'frete'],
            'senha_ravex': ['SENHA RAVEX', 'senha ravex', 'senha_ravex', 'ravex', 'senha'],
            'transporte': ['Nº Transporte', 'Transporte adicional', 'TRANSPORTE ORIGEM', 'transporte origem', 'transporte', 'transporte_de_origem', 'origem'],
            'cte_output': ['CTe gerado', 'Nº CTE', 'nº cte', 'numero cte', 'número cte', 'cte', 'n_cte', 'nºcte']
        }

        result = {}

        for key, names in mappings.items():
            index = self.find_column_by_name(names)
            if index != -1:
                result[key] = index

        return result
    
    def find_closest_column(self, search_term):
        """
        Encontra a coluna mais próxima usando similaridade de strings
        
        Args:
            search_term: Termo de busca
            
        Returns:
            Índice da coluna ou -1 se não encontrada
        """
        from difflib import SequenceMatcher
        
        search_normalized = search_term.lower().strip()
        best_match_idx = -1
        best_match_ratio = 0.5  # Mínimo 50% de similaridade
        
        for i, header in enumerate(self.headers):
            header_normalized = header.lower().strip()
            ratio = SequenceMatcher(None, search_normalized, header_normalized).ratio()
            
            if ratio > best_match_ratio:
                best_match_ratio = ratio
                best_match_idx = i
        
        return best_match_idx

    def get_preview_data(self, max_rows=10):
        """
        Obtém visualização formatada dos dados

        Args:
            max_rows: Número máximo de linhas

        Returns:
            Dicionário com dados de preview
        """
        return {
            'headers': self.headers,
            'rows': self.data[:max_rows],
            'total_rows': len(self.data),
            'total_columns': len(self.headers)
        }

    def get_cell_value(self, row_index, col_index):
        """
        Obtém valor de célula

        Args:
            row_index: Índice da linha
            col_index: Índice da coluna

        Returns:
            Valor da célula ou None
        """
        if 0 <= row_index < len(self.data):
            if 0 <= col_index < len(self.data[row_index]):
                return self.data[row_index][col_index]
        return None

    def set_cell_value(self, row_index, col_index, value):
        """
        Define valor de célula

        Args:
            row_index: Índice da linha
            col_index: Índice da coluna
            value: Novo valor
        """
        if row_index < 0:
            return

        # Garantir que a linha existe
        while len(self.data) <= row_index:
            self.data.append([])

        # Garantir que a coluna existe
        while len(self.data[row_index]) <= col_index:
            self.data[row_index].append('')

        self.data[row_index][col_index] = value
